# -*- coding: utf-8 -*-
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSX = Path(__file__).resolve().parents[2] / "LEMICHU_번개장터_상품데이터_34개.xlsx"


def as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def as_number(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default


def sheet_rows(workbook, name):
    worksheet = workbook[name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [as_text(cell) or f"col_{index}" for index, cell in enumerate(rows[0])]
    items = []
    for row in rows[1:]:
        item = {header: cell for header, cell in zip(headers, row)}
        items.append(item)
    return items


def image_urls_from_product(row):
    urls = []
    seen = set()
    for index in range(1, 15):
        url = as_text(row.get(f"이미지{index} URL"))
        if url and url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def image_urls_from_sheet(rows_by_id, product_id):
    urls = []
    seen = set()
    for row in sorted(rows_by_id.get(product_id, []), key=lambda item: as_number(item.get("이미지순서"), 99)):
        url = as_text(row.get("이미지URL"))
        if url and url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def parse(xlsx_path: Path):
    workbook = load_workbook(xlsx_path, data_only=True)
    products = sheet_rows(workbook, "상품데이터")
    images = sheet_rows(workbook, "이미지목록") if "이미지목록" in workbook.sheetnames else []
    images_by_id = {}
    for image in images:
        images_by_id.setdefault(as_text(image.get("원본상품ID")), []).append(image)

    payload = []
    for row in products:
        product_id = as_text(row.get("원본상품ID"))
        if not product_id:
            continue
        urls = image_urls_from_sheet(images_by_id, product_id) or image_urls_from_product(row)
        payload.append(
            {
                "sourceProductId": product_id,
                "name": as_text(row.get("상품명")),
                "brand": as_text(row.get("브랜드")),
                "brandEn": as_text(row.get("영문브랜드")),
                "categoryRaw": as_text(row.get("중분류")),
                "statusRaw": as_text(row.get("상품상태")),
                "gradeRaw": as_text(row.get("상태등급")),
                "colorRaw": as_text(row.get("색상")),
                "sizeRaw": as_text(row.get("사이즈")),
                "accessories": as_text(row.get("구성품")),
                "salePrice": as_number(row.get("판매가")),
                "stockQuantity": as_number(row.get("재고"), 1) or 1,
                "shippingType": as_text(row.get("배송유형")) or "국내배송",
                "detailSource": as_text(row.get("상품 상세설명 원본")),
                "keywords": as_text(row.get("검색키워드")),
                "sourceUrl": as_text(row.get("원본링크")),
                "images": urls,
            }
        )
    return payload


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")
    products = parse(xlsx_path)
    json.dump(products, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
