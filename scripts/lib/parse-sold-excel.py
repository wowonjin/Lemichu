# -*- coding: utf-8 -*-
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSX = Path(__file__).resolve().parents[2] / "레미츄_판매완료.xlsx"


def as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def as_number(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default


def image_urls(row):
    urls = []
    seen = set()
    for index in range(1, 16):
        url = as_text(row.get(f"image_url_{index}"))
        if url and url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def parse(xlsx_path: Path):
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [as_text(cell) or f"col_{index}" for index, cell in enumerate(rows[0])]
    payload = []
    for values in rows[1:]:
        row = {header: cell for header, cell in zip(headers, values)}
        product_id = as_text(row.get("product_id"))
        title = as_text(row.get("title"))
        images = image_urls(row)
        if not product_id or not title or not images:
            continue
        payload.append(
            {
                "sourceProductId": product_id,
                "statusRaw": as_text(row.get("status")),
                "name": title,
                "brand": as_text(row.get("brand")),
                "conditionRaw": as_text(row.get("product_condition")),
                "salePrice": as_number(row.get("price_krw")),
                "detailSource": as_text(row.get("description")),
                "sourceUrl": as_text(row.get("source_url")),
                "images": images,
            }
        )
    return payload


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")
    json.dump(parse(xlsx_path), sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
